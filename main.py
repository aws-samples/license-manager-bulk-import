from os.path import join, dirname, abspath
from openpyxl import Workbook, load_workbook
from registrars import product_registrars


statement_path = join(dirname(abspath(__file__)), 'Sample.xlsx')

statement = load_workbook(statement_path)
license_summary = statement['License Summary']


# Our license summary starts on row 5 - change the below if yours is different.
product_start_row = 5
# We are also assuming that there are 4 columns in the order of:
# category, product family, version, # of licenses
for row in license_summary.iter_rows(min_row=product_start_row, max_col=4):
    if row[1].value is not None:
        product_family = row[1].value
    product_version = row[2].value
    license_quantity = row[3].value

    # Look up the registration function for the product (see registrars.py)
    registrar = product_registrars.get(product_family, lambda product_family: print('Invalid product "{}"'.format(product_family)))
    registrar(family = product_family, version = product_version, quantity = license_quantity)
